import openpyxl as op
from openpyxl import Workbook , load_workbook
from openpyxl import get_coloums_letter
file = ("Enter File Name:")
book = load_workbook(file)
sheet = book.active
c=input("Enter the cell:")
#Modify the data in a cell 
sheet[c].value = input()
book.save(file)

# Name of all Sheets
def printAllsheetNames():
    print(book.sheetnames)
#Adding data to the sheet 
n = input("Enter number of rows")
for i in range(n):
    sheet.append([input(),input(),input()])
book.save(file)
#Access 
def AccessingCell(coloums,rows):
    coloums = int(input())
    rows = int(input())
    for row in range(0,coloums):
        for col in range(0,rows):
            char = get_coloums_letter(col)
            sheet[char+str(row)] = char + str(row)
    book.save(file)
#Merging cells
def MergeCells(c1,c2):
    sheet.merger_cells("'c1':'c2'")
    book.save(file)

